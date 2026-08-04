import io
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import backend.app as app_module
from backend.excel_export import build_catalogue_export


class ExportCursor:
    def __init__(self):
        self.rows = []

    def execute(self, query, params=None):
        if "FROM ingredients" in query:
            self.rows = [{"id": "tomato", "name": "番茄", "category": "蔬菜", "icon": "🍅"}]
        elif "FROM recipes" in query:
            self.rows = [{
                "id": "tomato-egg", "name": "番茄炒蛋", "description": "酸甜家常菜",
                "minutes": 10, "difficulty": "新手", "calories": 268, "tags": '["快手", "家常"]',
            }]
        elif "FROM recipe_ingredients" in query:
            self.rows = [{"recipe_id": "tomato-egg", "ingredient_id": "tomato", "position": 1}]
        elif "FROM recipe_seasonings" in query:
            self.rows = [{"recipe_id": "tomato-egg", "name": "食用油", "icon": "🫗", "position": 1}]
        elif "FROM recipe_steps" in query:
            self.rows = [{"recipe_id": "tomato-egg", "step_number": 1, "instruction": "切番茄"}]
        elif "FROM ingredient_categories" in query:
            self.rows = [{"name": "蔬菜"}]

    def fetchall(self):
        return self.rows

    def close(self):
        pass


class ExportConnection:
    def cursor(self, dictionary=False):
        return ExportCursor()

    def commit(self):
        pass

    def close(self):
        pass


class ExcelExportTests(unittest.TestCase):
    def test_export_has_import_compatible_sheets_and_rows(self):
        stream = build_catalogue_export(ExportConnection())
        workbook = load_workbook(stream, read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames, ["使用说明", "食材", "菜谱", "菜谱食材", "菜谱调味料", "步骤", "字典"])
        self.assertEqual(list(workbook.worksheets[1].values)[1], ("tomato", "番茄", "蔬菜", "🍅"))
        self.assertEqual(list(workbook.worksheets[2].values)[1][0], "tomato-egg")
        self.assertEqual(list(workbook.worksheets[3].values)[1], ("tomato-egg", "tomato", 1))
        self.assertEqual(list(workbook.worksheets[4].values)[1], ("tomato-egg", "食用油", "🫗", 1))
        self.assertEqual(list(workbook.worksheets[5].values)[1], ("tomato-egg", 1, "切番茄"))

    def test_export_endpoint_requires_admin_and_returns_attachment(self):
        client = app_module.app.test_client()
        self.assertEqual(client.get("/api/admin/export.xlsx").status_code, 401)
        with client.session_transaction() as session:
            session["admin_id"] = 1
            session["admin_username"] = "admin"
        with patch("backend.app.connect", return_value=ExportConnection()):
            response = client.get("/api/admin/export.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response.headers["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        self.assertEqual(workbook.worksheets[2].max_row, 2)


if __name__ == "__main__":
    unittest.main()
