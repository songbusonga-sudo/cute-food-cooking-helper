import io
import unittest
from unittest.mock import patch

from openpyxl import Workbook
from werkzeug.datastructures import FileStorage

from backend.excel_import import apply_import, read_upload, validate_payload
import backend.app as app_module


class FakeCursor:
    def __init__(self, ingredients=None, recipes=None):
        self.ingredients = ingredients or []
        self.recipes = recipes or []
        self.executed = []
        self.rows = []

    def execute(self, query, params=None):
        self.executed.append((query, params))
        if "FROM ingredients" in query:
            self.rows = self.ingredients
        elif "FROM recipes" in query:
            self.rows = self.recipes

    def executemany(self, query, params):
        self.executed.append((query, list(params)))

    def fetchall(self):
        return self.rows

    def close(self):
        pass


class FakeConnection:
    def __init__(self):
        self.cursor_instance = FakeCursor()
        self.committed = False
        self.rolled_back = False
        self.in_transaction = False

    def cursor(self, dictionary=False):
        return self.cursor_instance

    def start_transaction(self):
        if self.in_transaction:
            raise RuntimeError("Transaction already in progress")
        self.in_transaction = True

    def commit(self):
        self.committed = True
        self.in_transaction = False

    def rollback(self):
        self.rolled_back = True
        self.in_transaction = False

    def close(self):
        pass


class SeedCursor:
    def __init__(self, existing_recipe_ids):
        self.existing_recipe_ids = set(existing_recipe_ids)
        self.current_recipe_id = None
        self.executed = []

    def execute(self, query, params=None):
        self.executed.append((query, params))
        if "SELECT id FROM recipes" in query:
            self.current_recipe_id = params[0]

    def fetchone(self):
        return (self.current_recipe_id,) if self.current_recipe_id in self.existing_recipe_ids else None

    def close(self):
        pass


class SeedConnection:
    def __init__(self, existing_recipe_ids):
        self.cursor_instance = SeedCursor(existing_recipe_ids)
        self.committed = False

    def cursor(self, dictionary=False):
        return self.cursor_instance

    def commit(self):
        self.committed = True

    def close(self):
        pass


def make_upload(link_position=1):
    workbook = Workbook()
    ingredients = workbook.active
    ingredients.title = "食材"
    ingredients.append(["食材ID*", "食材名称*", "分类*", "图标*"])
    ingredients.append(["tomato", "番茄", "蔬菜", "🍅"])
    ingredients.append(["egg", "鸡蛋", "蛋奶", "🍳"])
    recipes = workbook.create_sheet("菜谱")
    recipes.append(["菜谱ID*", "菜谱名称*", "简介*", "用时(分钟)*", "难度*", "热量(kcal)*", "标签(用|分隔)*"])
    recipes.append(["tomato-egg", "番茄炒蛋", "酸甜下饭", 10, "新手", 268, "快手|家常"])
    links = workbook.create_sheet("菜谱食材")
    links.append(["菜谱ID*", "食材ID*", "顺序*"])
    links.append(["tomato-egg", "tomato", link_position])
    links.append(["tomato-egg", "egg", 2])
    steps = workbook.create_sheet("步骤")
    steps.append(["菜谱ID*", "步骤序号*", "做法*"])
    steps.append(["tomato-egg", 1, "切番茄并打散鸡蛋"])
    steps.append(["tomato-egg", 2, "翻炒至熟"])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return FileStorage(stream=stream, filename="menu.xlsx")


class ExcelImportTests(unittest.TestCase):
    def test_valid_workbook_generates_recipe_art_and_commits(self):
        data, upload_errors = read_upload(make_upload())
        self.assertEqual(upload_errors, [])
        connection = FakeConnection()
        errors, summary = validate_payload(data, connection, upload_errors)
        self.assertEqual(errors, [])
        self.assertEqual(data["recipes"][0]["art"], "🍅🍳")
        self.assertEqual(summary["recipes"]["insert"], 1)
        apply_import(data, connection)
        self.assertTrue(connection.committed)
        self.assertFalse(connection.rolled_back)

    def test_non_continuous_ingredient_order_is_rejected(self):
        data, upload_errors = read_upload(make_upload(link_position=2))
        errors, _ = validate_payload(data, FakeConnection(), upload_errors)
        self.assertTrue(any("食材顺序必须从 1 连续递增" in error["message"] for error in errors))

    def test_import_resets_validation_transaction_before_writing(self):
        data, upload_errors = read_upload(make_upload())
        connection = FakeConnection()
        connection.in_transaction = True
        errors, _ = validate_payload(data, connection, upload_errors)
        self.assertEqual(errors, [])
        apply_import(data, connection)
        self.assertTrue(connection.rolled_back)
        self.assertTrue(connection.committed)

    def test_custom_category_is_accepted_when_managed_by_backend(self):
        data, upload_errors = read_upload(make_upload())
        data["ingredients"][0]["分类*"] = "调味料"
        errors, _ = validate_payload(data, FakeConnection(), upload_errors, {"调味料", "蛋奶"})
        self.assertEqual(errors, [])

    def test_preview_and_commit_require_an_admin_session(self):
        app_module.IMPORT_PREVIEWS.clear()
        client = app_module.app.test_client()
        with client.session_transaction() as session:
            session["admin_id"] = 1
            session["admin_username"] = "admin"
        preview_connection = FakeConnection()
        commit_connection = FakeConnection()
        backup = {"id": 1, "file_name": "test-backup.xlsx", "summary": {"recipes": 0}}
        with patch("backend.app.connect", side_effect=[preview_connection, commit_connection]), patch(
            "backend.app.create_catalogue_backup", return_value=backup
        ), patch("backend.app.log_admin_action"):
            preview = client.post(
                "/api/admin/import/preview",
                data={"file": make_upload()},
                content_type="multipart/form-data",
            )
            self.assertEqual(preview.status_code, 200)
            payload = preview.get_json()
            self.assertTrue(payload["valid"])
            commit = client.post("/api/admin/import/commit", json={"token": payload["token"]})
        self.assertEqual(commit.status_code, 200)
        self.assertTrue(commit.get_json()["imported"])
        self.assertTrue(commit_connection.committed)

    def test_seed_catalogue_does_not_overwrite_an_existing_recipe(self):
        existing_recipe_id = app_module.RECIPES[0]["id"]
        connection = SeedConnection({existing_recipe_id})
        with patch("backend.app.connect", return_value=connection):
            app_module.seed_catalogue()
        recipe_inserts = [
            params[0]
            for query, params in connection.cursor_instance.executed
            if "INSERT INTO recipes" in query
        ]
        self.assertNotIn(existing_recipe_id, recipe_inserts)
        self.assertTrue(connection.committed)


if __name__ == "__main__":
    unittest.main()
