import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="4D8874")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MINT_FILL = PatternFill("solid", fgColor="E7F4EB")
CREAM_FILL = PatternFill("solid", fgColor="FFFAF0")
NOTE_FILL = PatternFill("solid", fgColor="FFF5D8")
LIGHT_BORDER = Border(bottom=Side(style="thin", color="FFF0E5D8"))
HEADER_BORDER = Border(bottom=Side(style="thin", color="FFECDCC8"))


def catalogue_export_data(connection):
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, name, category, icon FROM ingredients ORDER BY category, name")
        ingredients = cursor.fetchall()
        cursor.execute(
            """
            SELECT id, name, description, minutes, difficulty, calories, tags
            FROM recipes ORDER BY minutes, name
            """
        )
        recipes = cursor.fetchall()
        for recipe in recipes:
            try:
                recipe["tag_list"] = json.loads(recipe.get("tags") or "[]")
            except (TypeError, json.JSONDecodeError):
                recipe["tag_list"] = []
        cursor.execute(
            """
            SELECT recipe_id, ingredient_id, position
            FROM recipe_ingredients ORDER BY recipe_id, position
            """
        )
        links = cursor.fetchall()
        cursor.execute(
            """
            SELECT recipe_id, step_number, instruction
            FROM recipe_steps ORDER BY recipe_id, step_number
            """
        )
        steps = cursor.fetchall()
        cursor.execute("SELECT name FROM ingredient_categories ORDER BY name")
        categories = [row["name"] for row in cursor.fetchall()]
        return ingredients, recipes, links, steps, categories
    finally:
        cursor.close()


def style_header(sheet, row=1):
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    sheet.row_dimensions[row].height = 26
    sheet.freeze_panes = "A2"


def style_data(sheet, start_row, end_row):
    for row in sheet.iter_rows(min_row=start_row, max_row=max(start_row, end_row)):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column == 3)
            cell.border = LIGHT_BORDER


def build_catalogue_export(connection):
    ingredients, recipes, links, steps, categories = catalogue_export_data(connection)
    workbook = Workbook()
    guide = workbook.active
    guide.title = "使用说明"
    ingredient_sheet = workbook.create_sheet("食材")
    recipe_sheet = workbook.create_sheet("菜谱")
    link_sheet = workbook.create_sheet("菜谱食材")
    step_sheet = workbook.create_sheet("步骤")
    dictionary_sheet = workbook.create_sheet("字典")

    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:H1")
    guide["A1"] = "今天吃什么呀 - 当前数据库完整导出"
    guide["A1"].fill = HEADER_FILL
    guide["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    guide["A1"].alignment = Alignment(horizontal="center", vertical="center")
    guide.row_dimensions[1].height = 28
    guide_rows = [
        ("导出内容", "此文件由管理后台实时生成，包含当前食材、菜谱、食材关联、烹饪步骤和分类。"),
        ("再次导入", "请不要修改工作表名称和第一行表头；保存为 .xlsx 后可从“菜谱管理 - 导入 Excel”重新导入。"),
        ("菜谱图标", "无需填写。导入时后端会按“菜谱食材”中的顺序自动组合食材图标。"),
        ("分类", "“食材”里的分类必须已在管理后台“食材管理 - 分类管理”中创建。"),
        ("标签格式", "多个标签使用英文竖线 | 分隔，例如：快手|家常|清爽。"),
    ]
    for row_index, (label, value) in enumerate(guide_rows, start=3):
        guide.cell(row_index, 1, label).fill = MINT_FILL
        guide.cell(row_index, 1).font = Font(bold=True, color="4E4239")
        guide.cell(row_index, 2, value).fill = CREAM_FILL
        guide.cell(row_index, 2).alignment = Alignment(vertical="center", wrap_text=True)
        guide.row_dimensions[row_index].height = 28
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 78

    for sheet in (ingredient_sheet, recipe_sheet, link_sheet, step_sheet, dictionary_sheet):
        sheet.sheet_view.showGridLines = False

    ingredient_sheet.append(["食材ID*", "食材名称*", "分类*", "图标*"])
    for item in ingredients:
        ingredient_sheet.append([item["id"], item["name"], item["category"], item["icon"]])
    style_header(ingredient_sheet)
    style_data(ingredient_sheet, 2, len(ingredients) + 1)
    for column, width in {"A": 22, "B": 20, "C": 16, "D": 12}.items():
        ingredient_sheet.column_dimensions[column].width = width

    recipe_sheet.append(["菜谱ID*", "菜谱名称*", "简介*", "用时(分钟)*", "难度*", "热量(kcal)*", "标签(用|分隔)*"])
    for recipe in recipes:
        recipe_sheet.append([
            recipe["id"], recipe["name"], recipe["description"], recipe["minutes"],
            recipe["difficulty"], recipe["calories"], "|".join(recipe["tag_list"]),
        ])
    style_header(recipe_sheet)
    style_data(recipe_sheet, 2, len(recipes) + 1)
    for column, width in {"A": 24, "B": 22, "C": 52, "D": 14, "E": 14, "F": 14, "G": 28}.items():
        recipe_sheet.column_dimensions[column].width = width
    recipe_sheet.column_dimensions["C"].width = 52
    for row in range(2, len(recipes) + 2):
        recipe_sheet.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True)
    difficulty_validation = DataValidation(type="list", formula1="'字典'!$C$2:$C$4", allow_blank=False)
    recipe_sheet.add_data_validation(difficulty_validation)
    difficulty_validation.add(f"E2:E{max(2, len(recipes) + 1)}")

    link_sheet.append(["菜谱ID*", "食材ID*", "顺序*"])
    for link in links:
        link_sheet.append([link["recipe_id"], link["ingredient_id"], link["position"]])
    style_header(link_sheet)
    style_data(link_sheet, 2, len(links) + 1)
    for column, width in {"A": 24, "B": 24, "C": 12}.items():
        link_sheet.column_dimensions[column].width = width

    step_sheet.append(["菜谱ID*", "步骤序号*", "做法*"])
    for step in steps:
        step_sheet.append([step["recipe_id"], step["step_number"], step["instruction"]])
    style_header(step_sheet)
    style_data(step_sheet, 2, len(steps) + 1)
    for column, width in {"A": 24, "B": 14, "C": 78}.items():
        step_sheet.column_dimensions[column].width = width
    for row in range(2, len(steps) + 2):
        step_sheet.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True)

    dictionary_sheet.append(["食材分类", "", "菜谱难度"])
    max_rows = max(len(categories), 3)
    for index in range(max_rows):
        dictionary_sheet.append([
            categories[index] if index < len(categories) else None,
            None,
            ("新手", "简单", "进阶")[index] if index < 3 else None,
        ])
    style_header(dictionary_sheet)
    for row in dictionary_sheet.iter_rows(min_row=2, max_row=max_rows + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    dictionary_sheet.column_dimensions["A"].width = 18
    dictionary_sheet.column_dimensions["B"].width = 5
    dictionary_sheet.column_dimensions["C"].width = 18

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
