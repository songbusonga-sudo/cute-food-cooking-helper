import json
import re

from openpyxl import load_workbook


MAX_FILE_SIZE = 5 * 1024 * 1024
ID_PATTERN = re.compile(r"[a-z0-9-]{2,60}")
DEFAULT_CATEGORIES = {"蔬菜", "肉类", "蛋奶", "豆制品", "菌菇", "水产", "主食", "其他"}
DIFFICULTIES = {"新手", "简单", "进阶"}
SHEETS = {
    "食材": ["食材ID*", "食材名称*", "分类*", "图标*"],
    "菜谱": ["菜谱ID*", "菜谱名称*", "简介*", "用时(分钟)*", "难度*", "热量(kcal)*", "标签(用|分隔)*"],
    "菜谱食材": ["菜谱ID*", "食材ID*", "顺序*"],
    "菜谱调味料": ["菜谱ID*", "调味料名称*", "图标*", "顺序*"],
    "步骤": ["菜谱ID*", "步骤序号*", "做法*"],
}


def issue(sheet, row, field, message):
    return {"sheet": sheet, "row": row, "field": field, "message": message}


def text(value):
    return "" if value is None else str(value).strip()


def read_upload(file_storage):
    if not file_storage or not file_storage.filename:
        return None, [issue("文件", 0, "文件", "请选择要导入的 Excel 文件")]
    if not file_storage.filename.lower().endswith(".xlsx"):
        return None, [issue("文件", 0, "文件", "只支持 .xlsx 格式的模板文件")]

    stream = file_storage.stream
    stream.seek(0, 2)
    size = stream.tell()
    stream.seek(0)
    if size > MAX_FILE_SIZE:
        return None, [issue("文件", 0, "文件", "文件不能超过 5 MB")]

    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception:
        return None, [issue("文件", 0, "文件", "无法读取 Excel。请确认文件未损坏且不是受密码保护的文件")]

    errors = []
    for sheet_name in SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(issue(sheet_name, 1, "工作表", "缺少必需工作表"))

    if errors:
        workbook.close()
        return None, errors

    data = {"ingredients": [], "recipes": [], "links": [], "seasonings": [], "steps": []}
    targets = (
        ("食材", "ingredients"),
        ("菜谱", "recipes"),
        ("菜谱食材", "links"),
        ("菜谱调味料", "seasonings"),
        ("步骤", "steps"),
    )
    for sheet_name, target in targets:
        worksheet = workbook[sheet_name]
        expected = SHEETS[sheet_name]
        actual = [text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        while actual and not actual[-1]:
            actual.pop()
        if actual != expected:
            errors.append(issue(sheet_name, 1, "表头", f"表头必须为：{'、'.join(expected)}"))
            continue

        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(expected), values_only=True), start=2
        ):
            row = [text(value) for value in values]
            if not any(row):
                continue
            record = {header: row[index] for index, header in enumerate(expected)}
            record["_row"] = row_number
            for header, value in record.items():
                if header != "_row" and not value:
                    errors.append(issue(sheet_name, row_number, header, "该字段不能为空"))
            data[target].append(record)

    workbook.close()
    return data, errors


def as_integer(value, sheet, row, field, minimum, errors):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        errors.append(issue(sheet, row, field, "必须填写整数"))
        return None
    if parsed < minimum:
        errors.append(issue(sheet, row, field, f"必须大于等于 {minimum}"))
        return None
    return parsed


def validate_payload(data, connection, initial_errors=None, categories=None):
    errors = list(initial_errors or [])
    available_categories = set(categories or DEFAULT_CATEGORIES)
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT id, name, icon FROM ingredients")
    existing_ingredients = {row["id"]: row for row in cursor.fetchall()}
    cursor.execute("SELECT id, name FROM recipes")
    existing_recipes = {row["id"]: row for row in cursor.fetchall()}
    cursor.close()

    ingredient_ids = set()
    ingredient_names = set()
    input_ingredient_icons = {}
    for item in data["ingredients"]:
        row = item["_row"]
        ingredient_id = item["食材ID*"]
        name = item["食材名称*"]
        category = item["分类*"]
        icon = item["图标*"]
        if not ID_PATTERN.fullmatch(ingredient_id):
            errors.append(issue("食材", row, "食材ID*", "只能使用 2-60 位小写英文、数字和连字符"))
        if ingredient_id in ingredient_ids:
            errors.append(issue("食材", row, "食材ID*", "同一文件中不能重复"))
        ingredient_ids.add(ingredient_id)
        if name in ingredient_names:
            errors.append(issue("食材", row, "食材名称*", "同一文件中不能重复"))
        ingredient_names.add(name)
        if category not in available_categories:
            errors.append(issue("食材", row, "分类*", "该分类尚未在管理后台创建"))
        if len(name) > 40 or len(icon) > 16:
            errors.append(issue("食材", row, "名称或图标", "名称最长 40 字符，图标最长 16 字符"))
        input_ingredient_icons[ingredient_id] = icon

    existing_ingredient_names = {row["name"]: row["id"] for row in existing_ingredients.values()}
    for item in data["ingredients"]:
        existing_id = existing_ingredient_names.get(item["食材名称*"])
        if existing_id and existing_id != item["食材ID*"]:
            errors.append(issue("食材", item["_row"], "食材名称*", f"名称已被食材ID “{existing_id}” 使用"))

    recipe_ids = set()
    recipe_names = set()
    for item in data["recipes"]:
        row = item["_row"]
        recipe_id = item["菜谱ID*"]
        name = item["菜谱名称*"]
        if not ID_PATTERN.fullmatch(recipe_id):
            errors.append(issue("菜谱", row, "菜谱ID*", "只能使用 2-60 位小写英文、数字和连字符"))
        if recipe_id in recipe_ids:
            errors.append(issue("菜谱", row, "菜谱ID*", "同一文件中不能重复"))
        recipe_ids.add(recipe_id)
        if name in recipe_names:
            errors.append(issue("菜谱", row, "菜谱名称*", "同一文件中不能重复"))
        recipe_names.add(name)
        if len(name) > 80 or len(item["简介*"]) > 255:
            errors.append(issue("菜谱", row, "菜谱名称*或简介*", "名称最长 80 字符，简介最长 255 字符"))
        item["minutes"] = as_integer(item["用时(分钟)*"], "菜谱", row, "用时(分钟)*", 1, errors)
        item["calories"] = as_integer(item["热量(kcal)*"], "菜谱", row, "热量(kcal)*", 0, errors)
        if item["难度*"] not in DIFFICULTIES:
            errors.append(issue("菜谱", row, "难度*", "请选择模板“字典”页中的难度"))
        tags = [tag.strip() for tag in item["标签(用|分隔)*"].split("|")]
        if not tags or any(not tag for tag in tags):
            errors.append(issue("菜谱", row, "标签(用|分隔)*", "至少填写一个标签，多个标签用英文竖线 | 分隔"))
        item["tags"] = tags

    existing_recipe_names = {row["name"]: row["id"] for row in existing_recipes.values()}
    for item in data["recipes"]:
        existing_id = existing_recipe_names.get(item["菜谱名称*"])
        if existing_id and existing_id != item["菜谱ID*"]:
            errors.append(issue("菜谱", item["_row"], "菜谱名称*", f"名称已被菜谱ID “{existing_id}” 使用"))

    known_ingredients = set(existing_ingredients) | ingredient_ids
    links_by_recipe = {recipe_id: [] for recipe_id in recipe_ids}
    seen_links = set()
    seen_link_positions = set()
    for item in data["links"]:
        row = item["_row"]
        recipe_id = item["菜谱ID*"]
        ingredient_id = item["食材ID*"]
        position = as_integer(item["顺序*"], "菜谱食材", row, "顺序*", 1, errors)
        item["position"] = position
        if recipe_id not in recipe_ids:
            errors.append(issue("菜谱食材", row, "菜谱ID*", "必须是本次“菜谱”工作表中填写的菜谱ID"))
        if ingredient_id not in known_ingredients:
            errors.append(issue("菜谱食材", row, "食材ID*", "该食材ID不在本次文件，也不在数据库"))
        link_key = (recipe_id, ingredient_id)
        if link_key in seen_links:
            errors.append(issue("菜谱食材", row, "食材ID*", "同一道菜不能重复关联同一种食材"))
        seen_links.add(link_key)
        position_key = (recipe_id, position)
        if position is not None and position_key in seen_link_positions:
            errors.append(issue("菜谱食材", row, "顺序*", "同一道菜的食材顺序不能重复"))
        seen_link_positions.add(position_key)
        if recipe_id in links_by_recipe:
            links_by_recipe[recipe_id].append(item)

    seasonings_by_recipe = {recipe_id: [] for recipe_id in recipe_ids}
    seen_seasoning_positions = set()
    seen_seasoning_names = set()
    for item in data["seasonings"]:
        row = item["_row"]
        recipe_id = item["菜谱ID*"]
        name = item["调味料名称*"]
        icon = item["图标*"]
        position = as_integer(item["顺序*"], "菜谱调味料", row, "顺序*", 1, errors)
        item["position"] = position
        if recipe_id not in recipe_ids:
            errors.append(issue("菜谱调味料", row, "菜谱ID*", "必须是本次“菜谱”工作表中填写的菜谱ID"))
        if len(name) > 40 or len(icon) > 16:
            errors.append(issue("菜谱调味料", row, "调味料名称*或图标*", "名称最长 40 字符，图标最长 16 字符"))
        name_key = (recipe_id, name)
        if name_key in seen_seasoning_names:
            errors.append(issue("菜谱调味料", row, "调味料名称*", "同一道菜不能重复填写同一种调味料"))
        seen_seasoning_names.add(name_key)
        position_key = (recipe_id, position)
        if position is not None and position_key in seen_seasoning_positions:
            errors.append(issue("菜谱调味料", row, "顺序*", "同一道菜的调味料顺序不能重复"))
        seen_seasoning_positions.add(position_key)
        if recipe_id in seasonings_by_recipe:
            seasonings_by_recipe[recipe_id].append(item)

    steps_by_recipe = {recipe_id: [] for recipe_id in recipe_ids}
    seen_step_numbers = set()
    for item in data["steps"]:
        row = item["_row"]
        recipe_id = item["菜谱ID*"]
        step_number = as_integer(item["步骤序号*"], "步骤", row, "步骤序号*", 1, errors)
        item["step_number"] = step_number
        if recipe_id not in recipe_ids:
            errors.append(issue("步骤", row, "菜谱ID*", "必须是本次“菜谱”工作表中填写的菜谱ID"))
        key = (recipe_id, step_number)
        if step_number is not None and key in seen_step_numbers:
            errors.append(issue("步骤", row, "步骤序号*", "同一道菜的步骤序号不能重复"))
        seen_step_numbers.add(key)
        if recipe_id in steps_by_recipe:
            steps_by_recipe[recipe_id].append(item)

    all_icons = {ingredient_id: row["icon"] for ingredient_id, row in existing_ingredients.items()}
    all_icons.update(input_ingredient_icons)
    for recipe in data["recipes"]:
        recipe_id = recipe["菜谱ID*"]
        links = links_by_recipe[recipe_id]
        seasonings = seasonings_by_recipe[recipe_id]
        steps = steps_by_recipe[recipe_id]
        if not links:
            errors.append(issue("菜谱", recipe["_row"], "菜谱ID*", "该菜谱至少需要关联一种食材"))
        if not seasonings:
            errors.append(issue("菜谱", recipe["_row"], "菜谱ID*", "该菜谱至少需要填写一种调味料"))
        if not steps:
            errors.append(issue("菜谱", recipe["_row"], "菜谱ID*", "该菜谱至少需要填写一个步骤"))
        positions = sorted(item["position"] for item in links if item["position"] is not None)
        if positions and positions != list(range(1, len(positions) + 1)):
            errors.append(issue("菜谱食材", recipe["_row"], "顺序*", f"菜谱 “{recipe_id}” 的食材顺序必须从 1 连续递增"))
        step_numbers = sorted(item["step_number"] for item in steps if item["step_number"] is not None)
        if step_numbers and step_numbers != list(range(1, len(step_numbers) + 1)):
            errors.append(issue("步骤", recipe["_row"], "步骤序号*", f"菜谱 “{recipe_id}” 的步骤序号必须从 1 连续递增"))
        recipe["art"] = "".join(
            all_icons.get(item["食材ID*"], "") for item in sorted(links, key=lambda item: item["position"] or 0)
        )
        if len(recipe["art"]) > 32:
            errors.append(issue("菜谱", recipe["_row"], "菜谱食材", "组合后的食材图标过长，最多可保存 32 个字符"))

    summary = {
        "ingredients": {
            "insert": sum(item["食材ID*"] not in existing_ingredients for item in data["ingredients"]),
            "update": sum(item["食材ID*"] in existing_ingredients for item in data["ingredients"]),
        },
        "recipes": {
            "insert": sum(item["菜谱ID*"] not in existing_recipes for item in data["recipes"]),
            "update": sum(item["菜谱ID*"] in existing_recipes for item in data["recipes"]),
        },
        "links": len(data["links"]),
        "seasonings": len(data["seasonings"]),
        "steps": len(data["steps"]),
    }
    return errors, summary


def apply_import(data, connection):
    cursor = connection.cursor()
    try:
        # Validation queries may have opened MySQL's implicit read transaction.
        # Clear it before starting the single atomic import transaction.
        if getattr(connection, "in_transaction", False):
            connection.rollback()
        connection.start_transaction()
        cursor.executemany(
            """
            INSERT INTO ingredients (id, name, category, icon) VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE name=VALUES(name), category=VALUES(category), icon=VALUES(icon)
            """,
            [
                (item["食材ID*"], item["食材名称*"], item["分类*"], item["图标*"])
                for item in data["ingredients"]
            ],
        )
        cursor.executemany(
            """
            INSERT INTO recipes (id, name, art, description, minutes, difficulty, calories, tags)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE name=VALUES(name), art=VALUES(art), description=VALUES(description),
              minutes=VALUES(minutes), difficulty=VALUES(difficulty), calories=VALUES(calories), tags=VALUES(tags)
            """,
            [
                (
                    item["菜谱ID*"], item["菜谱名称*"], item["art"], item["简介*"], item["minutes"],
                    item["难度*"], item["calories"], json.dumps(item["tags"], ensure_ascii=False),
                )
                for item in data["recipes"]
            ],
        )
        recipe_ids = [item["菜谱ID*"] for item in data["recipes"]]
        cursor.executemany("DELETE FROM recipe_ingredients WHERE recipe_id=%s", [(recipe_id,) for recipe_id in recipe_ids])
        cursor.executemany("DELETE FROM recipe_seasonings WHERE recipe_id=%s", [(recipe_id,) for recipe_id in recipe_ids])
        cursor.executemany("DELETE FROM recipe_steps WHERE recipe_id=%s", [(recipe_id,) for recipe_id in recipe_ids])
        cursor.executemany(
            "INSERT INTO recipe_ingredients (recipe_id, ingredient_id, position) VALUES (%s, %s, %s)",
            [(item["菜谱ID*"], item["食材ID*"], item["position"]) for item in data["links"]],
        )
        cursor.executemany(
            "INSERT INTO recipe_seasonings (recipe_id, name, icon, position) VALUES (%s, %s, %s, %s)",
            [(item["菜谱ID*"], item["调味料名称*"], item["图标*"], item["position"]) for item in data["seasonings"]],
        )
        cursor.executemany(
            "INSERT INTO recipe_steps (recipe_id, step_number, instruction) VALUES (%s, %s, %s)",
            [(item["菜谱ID*"], item["step_number"], item["做法*"]) for item in data["steps"]],
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
